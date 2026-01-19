import os
import openpyxl

path = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))


def get_row_count(file_path, sheet_name):
    file = path + file_path
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_column_count(file_path, sheet_name):
    file = path + file_path
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_column


def get_row_index(file, sheet_name, row_id):
    for i in range(1, get_row_count(file, sheet_name) + 1):
        if read_data(file, sheet_name, i, 1) == row_id:
            return i


def read_data(file_path, sheet_name, row, column):
    file = path + file_path
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row, column).value
